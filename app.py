import csv
import openpyxl
from flask import Flask, request, jsonify, send_file
from tempfile import NamedTemporaryFile
import shutil
from datetime import datetime

app = Flask(__name__)


def modify_date_csv(file):
    temp_file = NamedTemporaryFile(mode='w', delete=False)
    with open(file, 'r') as input_file, temp_file:
        reader = csv.reader(input_file)
        writer = csv.writer(temp_file)

        for row in reader:
            row[1] = '/'.join(row[1].split('-')[::-1])  # Modify date column from 'dd-mm-yyyy' to 'mm/dd/yyyy'
            writer.writerow(row)

    shutil.move(temp_file.name, file)


def modify_date_xlsx(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    rows_to_write = []

    for column in sheet.iter_rows(min_row=2):
        code = column[25].value
        if code is None:
            description = column[24].value
            if check_diare(description):
                rows_to_write.append(column)
        elif check_string(code):
            rows_to_write.append(column)

    # Clear the existing sheet data
    sheet.delete_rows(1, sheet.max_row)

    # Iterate through the rows to write and write them to the sheet
    for row in rows_to_write:
        sheet.append(row)

    for column in sheet.iter_rows(min_row=1):
        tanggal = column[4]
        tanggal.value = modify_date_format(tanggal.value)

        # print('Tanggal Lahir ==> ', column[13].value)
        ttl = column[13]
        ttl.value = modify_date_format(ttl.value)

        code = column[25].value
        if check_string(code):
            rows_to_write.append(column)

        # date_parts = date_cell.value.split('-')
        # date_cell.value = '/'.join(date_parts[::-1])  # Modify date column from 'dd-mm-yyyy' to 'mm/dd/yyyy'

    workbook.save("output.xlsx")


def check_string(string):
    if string is not None:
        if 'A09' in string:
            return True
        elif ';' in string and 'A09' in string:
            return True
        else:
            return False


def check_diare(description):
    words_to_check = ["diare", "bab", "mencret"]
    if description is not None:
        for word in words_to_check:
            if word in description.lower():
                print("return true")
                return True


def modify_date_format(date_str):
    # Convert input date string to a datetime object
    date = datetime.strptime(date_str, '%d-%m-%Y')

    # Format the datetime object into the desired output format
    modified_date_str = datetime.strftime(date, '%m/%d/%Y')

    return modified_date_str


@app.route('/modify_date', methods=['POST'])
def modify_date():
    file = request.files['file']
    print('Opening file %r', file)
    if not file:
        return jsonify({'error': 'No file uploaded'})

    file_extension = file.filename.rsplit('.', 1)[1].lower()

    if file_extension == 'csv':
        file.save(file.filename)
        modify_date_csv(file.filename)
    elif file_extension == 'xlsx':
        file.save(file.filename)
        modify_date_xlsx(file.filename)
    else:
        return jsonify({'error': 'Invalid file format. Only CSV and XLSX are supported.'})

    return send_file(file.filename, as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True, port=3001)
